# 불량 없는 크기의 MLCC 이미지들의 크기를 검사하여 엑셀파일에 저장하는 코드
import cv2
import glob
from openpyxl import Workbook

def contour(src,path):
    gray = cv2.cvtColor(src, cv2.COLOR_BGR2GRAY)
    ret, electrode = cv2.threshold(gray, 127, 255, cv2.THRESH_BINARY)
    ret, all = cv2.threshold(gray, 70, 255, cv2.THRESH_BINARY)

    contours_all, hierarchy = cv2.findContours(all, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
    contours_electrode, hierarchy = cv2.findContours(electrode, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)

    area=[]
    area.append(path)
    for cnt in contours_all:
        if cv2.contourArea(cnt)>10000:
            print(cv2.contourArea(cnt))
            area.append(cv2.contourArea(cnt))
        
    for cnt in contours_electrode:
        if cv2.contourArea(cnt)>3000:
            print(cv2.contourArea(cnt))
            area.append(cv2.contourArea(cnt))
            
    print(area)
    return area

img_files = glob.glob('D:\MLCC_Image\P052012235019(NSW528)/*tif')

wb = Workbook()
ws = wb.active
ws['A1'] = '전체'
ws['B1'] = '전극1'
ws['C1'] = '전극2'

for path in img_files:
    print(path)
    img = cv2.imread(path)
    contour(img,path)
    ws.append(contour(img,path))
wb.save('C:\Final_project\TaeHyeon\MLCC_Area2.xlsx')