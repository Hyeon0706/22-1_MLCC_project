from openpyxl import load_Workbook
import pandas as pd



#df = pd.read_excel("C:\PythonDocuments\Excel\OkayImage.xlsx")
wb = load_Workbook("C:\PythonDocuments\Excel\OkayImage.xlsx")
ws = wb.active

data = []
cnt = 0
var = []

for _ in range(1,7):
    ws.cell(row=1, column=_, value = "(1," + str(_) +")")



'''
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        var.append(ws.cell(row=x, column=y).value)'''

ws['A1'] = 1
ws['A2'] = 2
ws['A3'] = 3
ws['B1'] = 1
ws['B2'] = 2
ws['B3'] = 3
