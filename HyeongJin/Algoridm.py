from asyncio.windows_events import NULL
from random import *
import os.path
from array import array
from tkinter import N
import cv2
import matplotlib.pyplot as plt
import sys  
import numpy as np
import PIL
import glob
from openpyxl import Workbook








targerdir = r"C:\PythonDocuments\TestImage1"

files = os.listdir(targerdir)

count = 0

hex_str1 = NULL
hex_str2 = NULL


for i in files:

    if os.path.isdir(targerdir + r"\\" + i):
        pass

    else :
        hex_str1 = i
        hex_str2 = str(hex_str2) + hex_str1
        #print("file : " + i)
        if i.count(".") == 1 : # . 이 한개일떄
            count += 1
            V = i.split(".")
        #    print("file Name : " + V[0])
        else :
            #print(len(i))
            for k in range(len(i)-1,0,-1): # . 이 여러개 일때
                if i[k] == ".":
                    print("file Name : "+i[:k])
                    break
#print(count)
#print(hex_str1)
#print(hex_str2)

#print("----------------------------------------------")                
new_hex_str2 = hex_str2[1:]
length = 9
hex_list = [new_hex_str2[0 + i:length + i] for i in range(0, len(new_hex_str2), length)]
#print(hex_list)
#print("----------------------------------------------")   

for i in range(len(hex_list)) :
    hex_list[i] = "C:\PythonDocuments\TestImage1\ " + hex_list[i]
    hex_list[i] = hex_list[i].replace(" ", "")
    
count = [0 for i in range(len(hex_list))]
count1 = [0 for i in range(len(hex_list))]

for i in range(len(hex_list)) :
    img = cv2.imread(hex_list[i])
    slice1 = img[88:171, 236:257]
    slice2 = img[78:170, 404:427]
    slice1 = cv2.resize(slice1, dsize=(200, 488), interpolation=cv2.INTER_AREA)
    slice2 = cv2.resize(slice2, dsize=(200, 488), interpolation=cv2.INTER_AREA)
    
    gray1 = cv2.cvtColor(slice1, cv2.COLOR_BGR2GRAY)    
    height1, width1 = gray1.shape
    gray2 = cv2.cvtColor(slice2, cv2.COLOR_BGR2GRAY)    
    height2, width2 = gray2.shape
    
    th = 255

    gray1 = cv2.inRange(gray1, th, 255)  
    gray2 = cv2.inRange(gray2, th, 255)  
    
    #cv2.imshow("img", slice1)
    #cv2.imshow("igra", gray)
    
    #cv2.waitKey()
    
    for iii in range(len(gray1)):
        for jjj in range(len(gray1[iii])):
            if gray1[iii][jjj] != 0:
                count[i] += 1
                
    for iii in range(len(gray2)):
        for jjj in range(len(gray2[iii])):
            if gray2[iii][jjj] != 0:
                count1[i] += 1            
                
    #print("카운트의 숫자'''''''''''''''''''''''''''''''''''''''''''")
    #print(count[i])
    
      
    

    

'''  
#이미지 출력
for i in range(3) :
    img = cv2.imread(hex_list[i])
    slice1 = img[88:171, 236:257]
    slice2 = img[78:170, 404:427]
    slice1 = cv2.resize(slice1, dsize=(200, 488), interpolation=cv2.INTER_AREA)
    slice2 = cv2.resize(slice2, dsize=(200, 488), interpolation=cv2.INTER_AREA)
    
    bluecount = 0;
    greencount = 0;
    redcount = 0;
    for iii in range(len(slice1)):
         for jjj in range(len(slice1[iii])):
            if slice1[iii][jjj][0] > 254:
                 bluecount += 1
            if slice1[iii][jjj][1] > 254:
                 greencount += 1
            if slice1[iii][jjj][2] > 254:
                redcount += 1       
    
    if img is None:
        sys.exit("Could not read the image.")
 
'''


#    print("총픽셀 수는 : ")
#    print(slice1.size/3)            
#    print("bluecount값은 : ")
#    print(bluecount)            
#    print("greencount값은 : ")
#    print(greencount)            
#    print("redcount값은 : ")
#    print(redcount)

str_left = [0] * len(hex_list)
str_right = [0] * len(hex_list)
  
for i in range(len(hex_list)) :
    str_left[i] = 'A' + str(i+2)
    str_right[i] = 'B' + str(i+2)


wb = Workbook()
ws = wb.active
ws['A1'] = 'ng255'
ws['B1'] = 'ng255'
'''
ws['C1'] = '전극2'
ws['D1'] = ''
'''
'''
for path in range(3):
    print(path)
    
    ws.append(count[0])
'''
for i in range(len(hex_list)) :
    ws[str_left[i]] = count[i]
    ws[str_right[i]] = count1[i]

wb.save('C:\PythonDocuments\Excel\OkayImage.xlsx')

    
'''
cv2.imshow("img", img)
cv2.imshow("slice1", slice1)
cv2.imshow("slice2", slice2)

    
cv2.waitKey()
cv2.destroyAllWindows()   
'''